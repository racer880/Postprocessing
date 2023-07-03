import os
import glob
import openpyxl
from PIL import Image
import tkinter as tk
from tkinter import filedialog


class ReadFotoMetadata:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Foto Metadaten lesen")
        self.excel_file_path = tk.StringVar()
        self.folder_path = tk.StringVar()
        self.photo_file_paths = []

        self.create_widgets()

    def create_widgets(self):
        # Excel-Datei auswählen
        excel_label = tk.Label(self.root, text="Excel-Datei:")
        excel_label.pack()

        excel_frame = tk.Frame(self.root)
        excel_frame.pack()

        excel_entry = tk.Entry(excel_frame, textvariable=self.excel_file_path, width=50)
        excel_entry.pack(side=tk.LEFT)

        excel_button = tk.Button(excel_frame, text="Durchsuchen", command=self.select_excel_file)
        excel_button.pack(side=tk.LEFT)

        # Ordner auswählen
        folder_label = tk.Label(self.root, text="Ordner:")
        folder_label.pack()

        folder_frame = tk.Frame(self.root)
        folder_frame.pack()

        folder_entry = tk.Entry(folder_frame, textvariable=self.folder_path, width=50)
        folder_entry.pack(side=tk.LEFT)

        folder_button = tk.Button(folder_frame, text="Durchsuchen", command=self.select_folder)
        folder_button.pack(side=tk.LEFT)

        # Metadaten auslesen und speichern
        read_button = tk.Button(self.root, text="Metadaten auslesen", command=self.read_metadata)
        read_button.pack()

        # Ergebnislabel
        self.result_label = tk.Label(self.root, text="")
        self.result_label.pack()

        self.root.mainloop()

    def select_excel_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel-Dateien", "*.xlsx *.xls")])
        self.excel_file_path.set(file_path)

    def select_folder(self):
        folder_path = filedialog.askdirectory()
        self.folder_path.set(folder_path)
        self.photo_file_paths = glob.glob(f"{folder_path}/*.jpg") + glob.glob(f"{folder_path}/*.jpeg") + glob.glob(f"{folder_path}/*.png")

    def read_metadata(self):
        excel_file_path = self.excel_file_path.get()
        folder_path = self.folder_path.get()

        if excel_file_path and folder_path:
            try:
                wb = openpyxl.load_workbook(excel_file_path)
                ws = wb.active

                # Füge die Überschriften hinzu, wenn die Excel-Datei leer ist
                if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
                    ws.cell(row=1, column=1, value="Dateiname")
                    ws.cell(row=1, column=2, value="Datum")
                    ws.cell(row=1, column=3, value="Kameramodell")
                    ws.cell(row=1, column=4, value="Dateigröße")
                    ws.cell(row=1, column=5, value="Belichtungszeit")
                    ws.cell(row=1, column=6, value="Verschlusszeit")
                    ws.cell(row=1, column=7, value="Brennweite")

                # Finde die nächste leere Zeile
                next_row = ws.max_row + 1

                duplicate_count = 0  # Anzahl der doppelten Dateien

                for photo_file_path in self.photo_file_paths:
                    image = Image.open(photo_file_path)
                    exif_data = image._getexif()

                    camera_model = exif_data.get(272, "-")
                    date_time = exif_data.get(306, "-")
                    file_size = os.path.getsize(photo_file_path)
                    exposure_time = exif_data.get(33434, "-")
                    shutter_speed = exif_data.get(37377, "-")
                    focal_length = exif_data.get(37386, "-")

                    file_title = os.path.basename(photo_file_path)

                    # Überprüfe, ob der Dateiname bereits vorhanden ist
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                        if row[0].value == file_title:
                            duplicate_count += 1
                            break
                    else:
                        # Schreibe die Metadaten in die entsprechenden Spalten
                        ws.cell(row=next_row, column=1, value=file_title)
                        ws.cell(row=next_row, column=2, value=date_time)
                        ws.cell(row=next_row, column=3, value=camera_model)
                        ws.cell(row=next_row, column=4, value=file_size)
                        ws.cell(row=next_row, column=5, value=str(exposure_time))
                        ws.cell(row=next_row, column=6, value=str(shutter_speed))
                        ws.cell(row=next_row, column=7, value=str(focal_length))
                        next_row += 1

                wb.save(excel_file_path)
                wb.close()

                if duplicate_count > 0:
                    self.result_label.config(text=f"Metadaten wurden erfolgreich in die Excel-Datei gespeichert.\n"
                                                  f"{duplicate_count} doppelte Dateien wurden übersprungen.")
                else:
                    self.result_label.config(text="Metadaten wurden erfolgreich in die Excel-Datei gespeichert.")

            except Exception as e:
                self.result_label.config(text="Fehler beim Speichern der Metadaten: " + str(e))

        else:
            self.result_label.config(text="Bitte wähle sowohl eine Excel-Datei als auch einen Ordner aus.")


app = ReadFotoMetadata()
