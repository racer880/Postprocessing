import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import matlab.engine
import subprocess


# Parameter
font_size_title = 24
font_size_text = 14
y_space = 15
font_name_title = "Arial bold"
font_name_text = "Arial"

sheet_names = []
sheet_names2 = []
sheet_names3 = []
path_browse_file = [3]
path_browse_file2 = [3]
path_browse_file3 = [3]
list_of_textfield = []
label_names = ["Template - Pfad",
               "Template - Name",
               "Pfad - Diagramm speichern",
               "Excel-Spalte 1 - Suchleiste",
               "Excel-Spalte 1 - Resultierender Pfad",
               "Excel-Spalte 2 - Suchleiste",
               "Excel-Spalte 2 - Resultierender Pfad",
               "Excel-Spalte 3 - Suchleiste",
               "Excel-Spalte 3 - Resultierender Pfad",
               "Diagrammtitel:",
               "Anfangs-/Endwert /Schrittweite X-Achse:",
               "Anfangs-/Endwert /Schrittweite Y-Achse:"
               ]
textfield_names = [r'C:\Users\patrick.grubert\Downloads\Git\PostProcessing\Postprocessing\App\Files',
                   'X_Achse_als_Referenz',
                   r'C:\Users\patrick.grubert\Downloads\Git\PostProcessing\Postprocessing\App\Files',
                   r' ',
                   "Beispieltitel",
                   -2.0,
                   2.0,
                   0.2,
                   -1,
                   1,
                   0.2]


def create_diagram(template_path, template_name, save_path, excel_path, excel_path2, excel_path3, title, b, c, d, e, f, g):
    if template_name == 'X_Achse_als_Referenz':
       # eng = matlab.engine.start_matlab(template_path)
        #eng.cd(template_path, nargout=0)
        #getattr(eng, template_name)(excel_path, nargout=0)
        # Pfad zum MATLAB-Skript
        matlab_script_path = r"C:\Users\patrick.grubert\Downloads\Git\PostProcessing\Postprocessing\App\Files"
        matlab_script_path_full = template_path + '\\' + template_name
        # Aufruf des MATLAB-Skripts
        subprocess.call(["matlab", "-nodesktop", "-nosplash", "-r", f"run('{matlab_script_path_full(excel_path)}'); input('Press Enter to exit');"])


class CreateDiagramMatlab(tk.Frame):
    def __init__(self, parent, master_class, buttons):
        tk.Frame.__init__(self, parent)
        self.sheet_name = tk.StringVar()
        self.sheet_name2 = tk.StringVar()
        self.sheet_name3 = tk.StringVar()
        self.parent = parent
        self.master_class = master_class
        self.config(bg='gray')

        self.construct_grid()
        self.create_gui_new(buttons)

    def construct_grid(self):
        # configure the grid
        for i in range(4):
            self.columnconfigure(i, weight=1)
        for j in range(15):
            self.rowconfigure(j, weight=1)

    def browse_file(self):
        global file_path, sheet_name
        file_path = filedialog.askopenfilename()
        try:
            workbook = openpyxl.load_workbook(file_path)
            path_browse_file[0] = file_path
            self.textfield3.delete(0, 255)
            self.textfield3.insert(0, file_path)
            self.textfield3.update()
            self.sheet_names = workbook.sheetnames
            self.sheet_name.set(self.sheet_names[0])
            self.sheet_option['menu'].delete(0, 'end')
            for name in self.sheet_names:
                self.sheet_option['menu'].add_command(label=name, command=tk._setit(self.sheet_name, name))
        except Exception as e:
            messagebox.showerror('Error', e)

    def browse_file2(self):
        global file_path2, sheet_name2
        file_path2 = filedialog.askopenfilename()
        try:
            workbook2 = openpyxl.load_workbook(file_path2)
            path_browse_file2[0] = file_path2
            self.textfield3_1.delete(0, 255)
            self.textfield3_1.insert(0, file_path2)
            self.textfield3_1.update()
            self.sheet_names2 = workbook2.sheetnames
            self.sheet_name2.set(self.sheet_names2[0])
            self.sheet_option2['menu'].delete(0, 'end')
            for name2 in self.sheet_names2:
                self.sheet_option2['menu'].add_command(label=name2, command=tk._setit(self.sheet_name2, name2))
        except Exception as e:
            messagebox.showerror('Error', e)

    def browse_file3(self):
        global file_path3, sheet_name3
        file_path3 = filedialog.askopenfilename()
        try:
            workbook3 = openpyxl.load_workbook(file_path3)
            path_browse_file3[0] = file_path3
            self.textfield3_2.delete(0, 255)
            self.textfield3_2.insert(0, file_path3)
            self.textfield3_2.update()
            self.sheet_names3 = workbook3.sheetnames
            self.sheet_name3.set(self.sheet_names3[0])
            self.sheet_option3['menu'].delete(0, 'end')
            for name3 in self.sheet_names3:
                self.sheet_option3['menu'].add_command(label=name3, command=tk._setit(self.sheet_name3, name3))
        except Exception as e:
            messagebox.showerror('Error', e)

    def choose_sheet(self, template_path, template_name):
        try:
            eng = matlab.engine.start_matlab(template_path)
            eng.cd(template_path, nargout=0)
            getattr(eng, template_name)(path_browse_file[0], nargout=0)

        except Exception as e:
            messagebox.showerror('Error', e)

    def create_gui_new(self, buttons):
        # Create a list of labels and textfield
        # Title
        self.title_label = tk.Label(self, text="Diagramme erstellen", font=(font_name_title, font_size_title),
                                    background="grey")
        self.title_label.grid(column=0, row=0, padx=5, pady=3 * y_space, sticky=tk.EW, columnspan=4)

        # Labels
        tk.Label(self, text=label_names[0], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=2, padx=2, pady=y_space)
        tk.Label(self, text=label_names[1], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=3, padx=2, pady=y_space)
        tk.Label(self, text=label_names[2], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=4, padx=2, pady=y_space)
        tk.Label(self, text=label_names[3], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=5, padx=2, pady=y_space)
        tk.Label(self, text=label_names[4], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=6, padx=2, pady=y_space)
        tk.Label(self, text=label_names[5], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=7, padx=2, pady=y_space)
        tk.Label(self, text=label_names[6], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=8, padx=2, pady=y_space)
        tk.Label(self, text=label_names[7], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=9, padx=2, pady=y_space)
        tk.Label(self, text=label_names[8], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=10, padx=2, pady=y_space)
        tk.Label(self, text=label_names[9], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=11, padx=2, pady=y_space)
        tk.Label(self, text=label_names[10], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=12, padx=2, pady=y_space)
        tk.Label(self, text=label_names[11], width=50, font=(font_size_text, font_size_text),
                 background="grey").grid(column=0, row=13, padx=2, pady=y_space)

        # Textfield 1
        self.textfield0 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield0.insert(0, textfield_names[0])
        self.textfield0.grid(column=1, row=2, padx=5, pady=y_space, columnspan=3, sticky=tk.EW)
        list_of_textfield.append(self.textfield0)

        # Textfield 2
        self.textfield1 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield1.insert(0, textfield_names[1])
        self.textfield1.grid(column=1, row=3, padx=5, pady=y_space, columnspan=3, sticky=tk.EW)
        list_of_textfield.append(self.textfield1)

        # Textfield 3
        self.textfield2 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield2.insert(0, textfield_names[2])
        self.textfield2.grid(column=1, row=4, padx=5, pady=y_space, columnspan=3, sticky=tk.EW)
        list_of_textfield.append(self.textfield2)

        # Excel-file 1
        self.browse_button = tk.Button(self, text="Suchen", command=lambda: self.browse_file())
        self.browse_button.grid(column=1, row=5, padx=5, pady=y_space, sticky=tk.EW, columnspan=2)
        # Excel-file 2
        self.browse_button2 = tk.Button(self, text="Suchen", command=lambda: self.browse_file2())
        self.browse_button2.grid(column=1, row=7, padx=5, pady=y_space, sticky=tk.EW, columnspan=2)
        # Excel-file 3
        self.browse_button3 = tk.Button(self, text="Suchen", command=lambda: self.browse_file3())
        self.browse_button3.grid(column=1, row=9, padx=5, pady=y_space, sticky=tk.EW, columnspan=2)

        # OptionMenu to select the sheet to use
        self.sheet_name.set('')
        self.sheet_option = tk.OptionMenu(self, self.sheet_name, '')
        self.sheet_option.grid(column=3, row=5, padx=5, pady=y_space, sticky=tk.EW, columnspan=1)
        # OptionMenu to select the sheet to use
        self.sheet_name2.set('')
        self.sheet_option2 = tk.OptionMenu(self, self.sheet_name2, '')
        self.sheet_option2.grid(column=3, row=7, padx=5, pady=y_space, sticky=tk.EW, columnspan=1)
        # OptionMenu to select the sheet to use
        self.sheet_name3.set('')
        self.sheet_option3 = tk.OptionMenu(self, self.sheet_name3, '')
        self.sheet_option3.grid(column=3, row=9, padx=5, pady=y_space, sticky=tk.EW, columnspan=1)

        # Textfield 4
        self.textfield3 = tk.Entry(self, font=(font_name_text, font_size_text), background="grey")
        self.textfield3.insert(0, textfield_names[3])
        self.textfield3.grid(column=1, row=6, padx=5, pady=y_space, columnspan=3, sticky=tk.EW)
        list_of_textfield.append(self.textfield3)
        # Textfield 4
        self.textfield3_1 = tk.Entry(self, font=(font_name_text, font_size_text), background="grey")
        self.textfield3_1.insert(0, textfield_names[3])
        self.textfield3_1.grid(column=1, row=8, padx=5, pady=y_space, columnspan=3, sticky=tk.EW)
        list_of_textfield.append(self.textfield3_1)
        # Textfield 4
        self.textfield3_2 = tk.Entry(self, font=(font_name_text, font_size_text), background="grey")
        self.textfield3_2.insert(0, textfield_names[3])
        self.textfield3_2.grid(column=1, row=10, padx=5, pady=y_space, columnspan=3, sticky=tk.EW)
        list_of_textfield.append(self.textfield3_2)

        # Textfield 5
        self.textfield4 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield4.insert(0, textfield_names[4])
        self.textfield4.grid(column=1, row=11, padx=5, pady=y_space, columnspan=3, sticky=tk.EW)
        list_of_textfield.append(self.textfield4)

        # Textfield 6
        self.textfield5 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield5.insert(0, textfield_names[5])
        self.textfield5.grid(column=1, row=12, padx=5, pady=y_space, columnspan=1, sticky=tk.EW)
        list_of_textfield.append(self.textfield5)

        # Textfield 7
        self.textfield6 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield6.insert(0, textfield_names[6])
        self.textfield6.grid(column=2, row=12, padx=5, pady=y_space, columnspan=1, sticky=tk.EW)
        list_of_textfield.append(self.textfield6)

        # Textfield 8
        self.textfield7 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield7.insert(0, textfield_names[7])
        self.textfield7.grid(column=3, row=12, padx=5, pady=y_space, columnspan=1, sticky=tk.EW)
        list_of_textfield.append(self.textfield7)

        # Textfield 9
        self.textfield8 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield8.insert(0, textfield_names[8])
        self.textfield8.grid(column=1, row=13, padx=5, pady=y_space, columnspan=1, sticky=tk.EW)
        list_of_textfield.append(self.textfield8)

        # Textfield 10
        self.textfield9 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield9.insert(0, textfield_names[9])
        self.textfield9.grid(column=2, row=13, padx=5, pady=y_space, columnspan=1, sticky=tk.EW)
        list_of_textfield.append(self.textfield9)

        # Textfield 11
        self.textfield10 = tk.Entry(self, font=(font_name_text, font_size_text))
        self.textfield10.insert(0, textfield_names[10])
        self.textfield10.grid(column=3, row=13, padx=5, pady=y_space, columnspan=1, sticky=tk.EW)
        list_of_textfield.append(self.textfield10)

        # Create folder button
        self.info_button = tk.Button(self, text="Erstellen", width=50, height=5, command=lambda: create_diagram(
            list_of_textfield[0].get(), list_of_textfield[1].get(), list_of_textfield[2].get(),
            list_of_textfield[3].get(), list_of_textfield[4].get(), list_of_textfield[5].get(),
            list_of_textfield[6].get(), float(list_of_textfield[7].get()), float(list_of_textfield[8].get()),
            float(list_of_textfield[9].get()), float(list_of_textfield[10].get()), float(list_of_textfield[11].get())
            , float(list_of_textfield[12].get())))
        self.info_button.grid(column=1, row=15, padx=5, pady=2*y_space, sticky=tk.NS)

        # Back to Homepage
        self.back_button = tk.Button(self, text="Zurück", width=50, height=5,
                                     command=lambda: self.back_to_homepage(buttons))
        self.back_button.grid(column=2, row=15, padx=5, pady=2*y_space, sticky=tk.NS, columnspan=1)

        # TEST - Call compare function
        self.test_button = tk.Button(self, text="Compare", width=50, height=5, command=lambda: self.choose_sheet(
            list_of_textfield[0].get(), list_of_textfield[1].get()))
        self.test_button.grid(column=3, row=15, padx=5, pady=2*y_space, sticky=tk.NS, columnspan=1)

    def back_to_homepage(self, buttons):
        self.destroy()
        # Create a label for the title
        self.master.title_label.grid(row=0, column=1, columnspan=4, pady=40)
        k = 0
        for i in range(4):
            for j in range(4):
                buttons[k].grid(row=i + 1, column=j + 1, padx=20, pady=20)
                k = k + 1
