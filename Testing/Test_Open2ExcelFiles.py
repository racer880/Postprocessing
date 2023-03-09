import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt

# Function to browse for the Excel file and sheet to use
def browse_file():
    global file_path, sheet_name
    file_path = filedialog.askopenfilename()
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
        sheet_name.set(sheet_names[0])
        sheet_option['menu'].delete(0, 'end')
        for name in sheet_names:
            sheet_option['menu'].add_command(label=name, command=tk._setit(sheet_name, name))
    except Exception as e:
        messagebox.showerror('Error', e)

# Function to browse for the second Excel file and sheet to use
def browse_file2():
    global file_path2, sheet_name2
    file_path2 = filedialog.askopenfilename()
    try:
        workbook = openpyxl.load_workbook(file_path2)
        sheet_names = workbook.sheetnames
        sheet_name2.set(sheet_names[0])
        sheet_option2['menu'].delete(0, 'end')
        for name in sheet_names:
            sheet_option2['menu'].add_command(label=name, command=tk._setit(sheet_name2, name))
    except Exception as e:
        messagebox.showerror('Error', e)

# Function to compare two columns from two different Excel sheets
def compare_columns():
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name.get())
        df2 = pd.read_excel(file_path2, sheet_name=sheet_name2.get())
        col1 = col1_entry.get()
        col2 = col2_entry.get()
        data1 = df[col1].tolist()
        data2 = df2[col2].tolist()
        fig, ax = plt.subplots()
        ax.plot(data1, label='Sheet 1')
        ax.plot(data2, label='Sheet 2')
        ax.legend()
        ax.set_xlabel('Index')
        ax.set_ylabel('Values')
        ax.set_title(f'Comparison of {col1} and {col2}')
        scaling = float(scaling_entry.get())
        fig.set_size_inches(scaling, scaling)
        eps_path = filedialog.asksaveasfilename(defaultextension='.eps')
        fig.savefig(eps_path, format='eps')
        plt.close()
    except Exception as e:
        messagebox.showerror('Error', e)

# Create the GUI window
root = tk.Tk()
root.title('Excel Column Comparison')
root.geometry("1920x1080")

# Browse button for the Excel file and sheet to use
file_button = tk.Button(root, text='Browse for Excel File', command=browse_file)
file_button.pack(pady=10)

# OptionMenu to select the sheet to use
sheet_name = tk.StringVar()
sheet_name.set('')
sheet_option = tk.OptionMenu(root, sheet_name, '')
sheet_option.pack(pady=10)

# Browse button for the second Excel file
file2_button = tk.Button(root, text='Browse for Second Excel File', command=browse_file2)
file2_button.pack(pady=10)

# OptionMenu to select the sheet of the second Excel file to use
sheet_name2 = tk.StringVar()
sheet_name2.set('')
sheet_option2 = tk.OptionMenu(root, sheet_name2, '')
sheet_option2.pack(pady=10)

# Entry for the first column name
col1_label = tk.Label(root, text='Enter the name of the first column:')
col1_label.pack()
col1_entry = tk.Entry(root)
col1_entry.pack()

# Entry for the second column name
col2_label = tk.Label(root, text='Enter the name of the second column:')
col2_label.pack()
col2_entry = tk.Entry(root)
col2_entry.pack()

# Entry for the scaling factor
scaling_label = tk.Label(root, text='Enter the scaling factor for the exported EPS file:')
scaling_label.pack()
scaling_entry = tk.Entry(root)
scaling_entry.pack()

# Button to compare the columns and export the EPS file
compare_button = tk.Button(root, text='Compare Columns and Export EPS', command=compare_columns)
compare_button.pack(pady=20)

# Start the GUI event loop
root.mainloop()