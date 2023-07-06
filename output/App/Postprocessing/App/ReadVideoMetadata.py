import os
import glob
import openpyxl
import tkinter as tk
from tkinter import filedialog
from moviepy.editor import VideoFileClip

# Parameter
font_size_title = 24
font_size_text = 14
y_space = 20
font_name_title = "Arial bold"
font_name_text = "Arial"
sheet_names = []

class ReadVideoMetadata(tk.Frame):
    def __init__(self, parent, master_class, buttons):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.master_class = master_class
        self.config(bg='gray')
        self.excel_file_path = tk.StringVar()
        self.folder_path = tk.StringVar()
        self.video_file_paths = []

        self.construct_grid()
        self.create_widgets(buttons)


    def construct_grid(self):
        # configure the grid
        for i in range(10):
            self.columnconfigure(i, weight=1)
        for j in range(10):
            self.rowconfigure(j, weight=1)

    def create_widgets(self, buttons):

        self.title_label = tk.Label(self, text="Metadaten von Videos auslesen", font=(font_name_title, font_size_title),background="grey")
        self.title_label.grid(column=0, row=0, padx=520, pady=2 * y_space, sticky=tk.NSEW, columnspan=6)

        # Excel-Datei auswählen
        self.excel_label = tk.Label(self, text="Excel-Datei:",  background="grey")
        self.excel_label.grid(column=0, row=1, padx=5, pady=2 * y_space, sticky=tk.NSEW)
        self.excel_entry = tk.Entry(self, textvariable=self.excel_file_path, width=50)
        self.excel_entry.grid(column=1, row=1, padx=5, pady=2 * y_space, sticky=tk.NSEW, columnspan=3)
        self.excel_button = tk.Button(self, text="Durchsuchen", command=self.select_excel_file)
        self.excel_button.grid(column=4, row=1, padx=5, pady=2 * y_space, sticky=tk.NSEW)

        # Ordner auswählen
        self.folder_label = tk.Label(self, text="Ordner:", background="grey")
        self.folder_label.grid(column=0, row=2, padx=5, pady=2 * y_space, sticky=tk.NSEW)
        self.folder_entry = tk.Entry(self, textvariable=self.folder_path, width=50)
        self.folder_entry.grid(column=1, row=2, padx=5, pady=2 * y_space, sticky=tk.NSEW, columnspan=3)
        self. folder_button = tk.Button(self, text="Durchsuchen", command=self.select_folder)
        self.folder_button.grid(column=4, row=2, padx=5, pady=2 * y_space, sticky=tk.NSEW)

        # Metadaten auslesen und speichern
        self.read_button = tk.Button(self, text="Metadaten auslesen", width=50,height=5, command=self.read_metadata)
        self.read_button.grid(column=1, row=3, padx=5, pady=2 * y_space, sticky=tk.EW)

        # Back to Homepage
        self.info_button = tk.Button(self, text="Zurück", width=50, height=5, command=lambda: self.back_to_homepage(buttons))
        self.info_button.grid(column=3, row=3, padx=5, pady=2*y_space, sticky=tk.EW)

        # Ergebnislabel
        self.result_label2 = tk.Label(self, text="Ergebnis:", background="grey")
        self.result_label2.grid(column=0, row=4, padx=5, pady=2 * y_space, sticky=tk.NSEW)
        self.result_label = tk.Label(self, text="", background="grey")
        self.result_label.grid(column=1, row=4, padx=5, pady=2 * y_space, sticky=tk.NSEW, columnspan=6)

        # Details
        self.infobox_label = tk.Label(self, text="________________________________________",
                                      font=(font_name_title, font_size_title), background="grey")
        self.infobox_label.grid(column=0, row=5, padx=0, pady=2 * y_space, sticky=tk.EW, columnspan=8)
        self.infobox2_label = tk.Label(self, text="Details zur App:", font=(font_name_title, font_size_text),
                                       background="grey")

        self.infobox2_label.grid(column=0, row=6, padx=50, pady=y_space, sticky=tk.W, columnspan=8)
        self.infobox3_label = tk.Label(self, text=" - Ziel der App: Metadaten von Videos auslesen", font=(font_name_title, font_size_text), background="grey")
        self.infobox3_label.grid(column=0, row=7, padx=50, pady= y_space, sticky=tk.W, columnspan=8)

    def select_excel_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel-Dateien", "*.xlsx *.xls")])
        self.excel_file_path.set(file_path)

    def select_folder(self):
        folder_path = filedialog.askdirectory()
        self.folder_path.set(folder_path)
        self.video_file_paths = glob.glob(f"{folder_path}/*.mp4") + glob.glob(f"{folder_path}/*.mov")

    def read_metadata(self):
        excel_file_path = self.excel_file_path.get()
        folder_path = self.folder_path.get()

        if excel_file_path and folder_path:
            try:
                wb = openpyxl.load_workbook(excel_file_path)
                ws = wb.active

                # Füge die Überschriften hinzu, wenn die Excel-Datei leer ist
                if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
                    ws.cell(row=1, column=1, value="Dateiname")
                    ws.cell(row=1, column=2, value="Dauer")
                    ws.cell(row=1, column=3, value="Bildgröße")
                    ws.cell(row=1, column=4, value="Codec")
                    ws.cell(row=1, column=5, value="Bildrate")
                    ws.cell(row=1, column=6, value="Datum")

                # Finde die nächste leere Zeile
                next_row = ws.max_row + 1

                duplicate_count = 0

                for video_file_path in self.video_file_paths:
                    file_title = os.path.basename(video_file_path)

                    # Überprüfe, ob der Dateiname bereits vorhanden ist
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                        if row[0].value == file_title:
                            duplicate_count += 1
                            break
                    else:
                        try:
                            video = VideoFileClip(video_file_path)

                            duration = video.duration
                            size = os.path.getsize(video_file_path)
                            codec = video.reader.infos.get('video_codec_name', "-")
                            frame_rate = video.fps if hasattr(video, 'fps') else "-"

                            # Extrahiere das Erstellungsdatum aus dem Dateinamen
                            creation_date = self.extract_creation_date(file_title)

                            # Schreibe die Metadaten in die entsprechenden Spalten
                            ws.cell(row=next_row, column=1, value=file_title)
                            ws.cell(row=next_row, column=2, value=duration)
                            ws.cell(row=next_row, column=3, value=size)
                            ws.cell(row=next_row, column=4, value=codec)
                            ws.cell(row=next_row, column=5, value=frame_rate)
                            ws.cell(row=next_row, column=6, value=creation_date)
                            next_row += 1

                        except Exception as e:
                            ws.cell(row=next_row, column=1, value=file_title)
                            ws.cell(row=next_row, column=2, value="-")
                            ws.cell(row=next_row, column=3, value="-")
                            ws.cell(row=next_row, column=4, value="-")
                            ws.cell(row=next_row, column=5, value="-")
                            ws.cell(row=next_row, column=6, value="-")
                            next_row += 1

                wb.save(excel_file_path)
                wb.close()

                if duplicate_count > 0:
                    self.result_label.config(
                        text=f"Metadaten wurden erfolgreich in die Excel-Datei gespeichert.\n"
                             f"{duplicate_count} doppelte Dateien wurden übersprungen.")
                else:
                    self.result_label.config(text="Metadaten wurden erfolgreich in die Excel-Datei gespeichert.")

            except Exception as e:
                self.result_label.config(text="Fehler beim Speichern der Metadaten: " + str(e))

        else:
            self.result_label.config(text="Bitte wähle sowohl eine Excel-Datei als auch einen Ordner aus.")

    @staticmethod
    def extract_creation_date(file_name):
        # Implementiere die Logik zum Extrahieren des Erstellungsdatums aus dem Dateinamen
        # Hier ein Beispiel, wie das Erstellungsdatum aus dem Format "YYYYMMDD_HHMMSS" extrahiert werden könnte
        date_str = file_name.split("_")[0]
        if len(date_str) == 8:
            year = date_str[:4]
            month = date_str[4:6]
            day = date_str[6:8]
            return f"{day}.{month}.{year}"
        else:
            return "-"

    def back_to_homepage(self, buttons):
        self.destroy()
        # Create a label for the title
        self.master.title_label.grid(row=0, column=1, columnspan=4, pady=40)
        k = 0
        for i in range(4):
            for j in range(4):
                buttons[k].grid(row=i + 1, column=j + 1, padx=20, pady=20)
                k = k + 1

